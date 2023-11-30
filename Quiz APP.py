from tkinter import *
from tkinter import filedialog as fd
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import messagebox
import random
import re
import time

root = Tk()
root.title("Quiz App")
root.geometry("400x350")

# Globals
# File path of test spreadsheet
test_filename = "C://Users/USER/Documents/PYTHON PROJECTS/QUIZ GUI APP/TESTS/Test.xlsx"
test = {}  # {Question:[option a, option b, option c, option d]}
questions = []  # List of questions
index = 0  # Index of questions in questions list
score = 0  # Participant's score
workbook = load_workbook(filename=test_filename)  # Loads test spreadsheet
active_sheet = workbook.active  # Instantiates an active sheet object
# Dictionary to collect the information of the student given in the onboarding form
student_info = {}
choices = {}  # Dictionary to store the choices of the candidate

# =======================================Functions==========================================

# -----------------------------------------When Submit button is pressed
def turn_in():
    global title_label
    global q_label
    global q_frame
    global nxt_btn
    global prv_btn
    global student_info
    global submit_btn
    global questions
    global index
    global score
    global choices
    global var

    # Updates the choices dictionary with the very last choice provided before the submit button was pressed
    choices[questions[index]] = var.get()

    # Asks to be sure that candidate wants to turn in quiz
    submit_test = True
    # If candidate says OK
    if submit_test:
        # Remove Questions and Title
        q_frame.grid_forget()
        title_label.grid_forget()
        time_label.grid_forget()

        # Iterate through the choices dictionary items and check if the choice is the same as the answer which is last element of options list
        for question, choice in choices.items():
            if choice == test[question][-1]:
                # Update score
                score += 1

        # Display candidate details and candidate's score
        score_frame = LabelFrame(root, text="Final Score")
        score_frame.grid(row=0, column=0, padx=50, pady=50)
        details_label = Label(
            score_frame, text="Name: " + student_info["name"] + "\nLevel: " + student_info["level"] + "\nMatric: " + student_info["matric"] + "\n----------------" + "\nScore: " + str(score) + "/" + str(len(questions)))
        details_label.grid(row=0, column=0, ipadx=40, ipady=40)

        quit_btn = Button(root, text="End Quiz",
                          command=root.destroy, bg="#2e6930", fg="#fff")
        quit_btn.grid(row=1, column=0)

def submit():
    global title_label
    global time_label
    global q_label
    global q_frame
    global nxt_btn
    global prv_btn
    global student_info
    global submit_btn
    global questions
    global index
    global score
    global choices
    global var

    # Updates the choices dictionary with the very last choice provided before the submit button was pressed
    choices[questions[index]] = var.get()

    # Asks to be sure that candidate wants to turn in quiz
    submit_test = messagebox.askokcancel(
        title="Submit Quiz", message="Do you wish to submit? Cross-check your answers ")
    # If candidate says OK
    if submit_test:
        # Remove Questions and Title
        q_frame.grid_forget()
        title_label.grid_forget()
        time_label.grid_forget()

        # Iterate through the choices dictionary items and check if the choice is the same as the answer which is last element of options list
        for question, choice in choices.items():
            if choice == test[question][-1]:
                # Update score
                score += 1

        # Display candidate details and candidate's score
        score_frame = LabelFrame(root, text="Final Score")
        score_frame.grid(row=0, column=0, padx=50, pady=50)
        details_label = Label(
            score_frame, text="Name: " + student_info["name"] + "\nLevel: " + student_info["level"] + "\nMatric: " + student_info["matric"] + "\n----------------" + "\nScore: " + str(score) + "/" + str(len(questions)))
        details_label.grid(row=0, column=0, ipadx=40, ipady=40)

        quit_btn = Button(root, text="End Quiz",
                          command=root.destroy, bg="#2e6930", fg="#fff")
        quit_btn.grid(row=1, column=0)

#--------------------------------------------------------------Starts timer when take test is pressed
def run_timer():

    global active_sheet
    global time_label
    global q_frame

    #Create Variables for each of the units/hands: Hour, Minute, Seconds
    hourString = StringVar()
    minuteString = StringVar()
    secondString = StringVar()

    #Go to Time Field in Spreadsheet and pick the time string, then use regex to isolate the string digits:       ['00', '00', '00']
    time_string = active_sheet["B2"].value
    time_digits = re.findall(r'\d+', time_string)

    #Set each of these time string variables to these strings
    hourString.set(time_digits[0])
    minuteString.set(time_digits[1])
    secondString.set(time_digits[2])

    #Create Labels for them on the screen Row = 3, Column = 1
    time_label = Label(root, text="Time: " + hourString.get() +
                       ": " + minuteString.get() + ": " + secondString.get())
    time_label.grid(row=3, column=1)

    clockTime = int(hourString.get()) * 3600 + \
        int(minuteString.get()) * 60 + int(secondString.get())

    while clockTime > -1:

        totalMinutes, totalSeconds = divmod(clockTime, 60)
        totalHours = 0

        if totalMinutes > 60:

            totalHours, totalMinutes = divmod(totalMinutes, 60)

        hourString.set("{0:2d}".format(totalHours))
        minuteString.set("{0:2d}".format(totalMinutes))
        secondString.set("{0:2d}".format(totalSeconds))

        time_label = Label(root, text="Time: " + hourString.get() +
                           ": " + minuteString.get() + ": " + secondString.get())
        time_label.grid(row=3, column=1)

        root.update()
        time.sleep(1)

        #Let User know when time elapses
        if (clockTime == 0):
          turn_in()
          messagebox.showinfo("Quiz Ended", "Your time has expired")
          
        clockTime -= 1

# -----------------------------------When previous button is pressed
def prv_que():
    global index
    global questions
    global nxt_btn
    global prv_btn
    global submit_btn
    global q_frame
    global choices
    global var

    # Update the choices dictionary
    choices[questions[index]] = var.get()

    # Goes to previous question by reducing index by 1
    index -= 1

    # Displays questions and options
    display_questions(index)
    display_options(questions[index])
    # Checks if index is greater than 0, to check that the participant isnt on first question, and keeps the previous button normal
    if index > 0:
        prv_btn = Button(q_frame, text="Previous",
                         command=prv_que, bg="#2e6930", fg="#fff")
        prv_btn.grid(row=3, column=0, padx=10, pady=10)
        nxt_btn = Button(q_frame, text="Next", command=nxt_que,
                         state=NORMAL, bg="#2e6930", fg="#fff")
        nxt_btn.grid(row=3, column=2, padx=10, pady=10)
        if submit_btn:
            submit_btn.grid_forget()
    # However, if participant is still on first question, previous button is disabled
    elif index == 0:
        prv_btn = Button(q_frame, text="Previous",
                         command=prv_que, state=DISABLED, bg="#2e6930", fg="#fff")
        prv_btn.grid(row=3, column=0, padx=10, pady=10)
        nxt_btn = Button(q_frame, text="Next", command=nxt_que,
                         state=NORMAL, bg="#2e6930", fg="#fff")
        nxt_btn.grid(row=3, column=2, padx=10, pady=10)

# -------------------------------------When Next button is clicked


def nxt_que():
    global index
    global questions
    global nxt_btn
    global prv_btn
    global submit_btn
    global q_frame
    global choices
    global var

    # Update the choices dictionary
    choices[questions[index]] = var.get()

    # Move to next question by increasing index by 1
    index += 1
    # Display questions and options
    display_questions(index)
    display_options(questions[index])
    # If the participant is at the second to the last question or any question before, the next button will be normal and previous button will also switch to normal from the initial disabled stat
    if index < len(questions)-1:
        prv_btn = Button(q_frame, text="Previous",
                         command=prv_que, bg="#2e6930", fg="#fff")
        prv_btn.grid(row=3, column=0, padx=10, pady=10)
        nxt_btn = Button(q_frame, text="Next", command=nxt_que,
                         state=NORMAL, bg="#2e6930", fg="#fff")
        nxt_btn.grid(row=3, column=2, padx=10, pady=10)
    # However, when they arrive at the last question, the next button turns into a submit button
    else:
        nxt_btn.grid_forget()
        submit_btn = Button(q_frame, text="Submit",
                            command=submit, bg="#2e6930", fg="#fff")
        submit_btn.grid(row=3, column=2, padx=10, pady=10)
        prv_btn = Button(q_frame, text="Previous",
                         command=prv_que, bg="#2e6930", fg="#fff")
        prv_btn.grid(row=3, column=0, padx=10, pady=10)

# ----------------------------------Function executes when take test button is pressed


def take_test():
    global onboard
    global title_label
    global test
    global questions
    global name_field
    global matric_field
    global level_field
    global q_frame
    global q_label
    global prv_btn
    global nxt_btn
    global index
    global active_sheet
    global student_info
    global choices

    # Filling up the dictionary with these details form the fields
    student_info["name"] = name_field.get()
    student_info["level"] = level_field.get()
    student_info["matric"] = matric_field.get()

    # A generator that yields the serial numbers on column 1 of the active sheet
    def track_row():
        for row in active_sheet.iter_rows(min_row=4, max_col=1, values_only=True):
            yield row[0]

    # Iteration through each yield of the generator to ascertain the actual rows of the questions and options we want to use, The questions and options are then fixed into the test dictionary at the very beginning of this script
    for num in track_row():
        row = num + 3
        question = active_sheet["B" + str(row)].value
        options = []
        for row in active_sheet.iter_rows(min_row=row, max_row=row, min_col=3, values_only=True):
            for option in row:
                # This is a validation checkpoint for the integer valued options such as the years. COnverts them to strings
                if type(option) == int:
                    options.append(str(option))
                else:
                    options.append(option)
        # Adds the question and options list to the dictionary, as a key-value pair
        test[question] = options

    # Creates a list of questions from the keys of the test dictionary
    questions = list(test.keys())
    # Shuffles up the question so that the orders are different
    random.shuffle(questions)

    # Set each question to an initial choice of 0 in the choices dictionary, so that the radiobuttons highlight nothing until that choice is changed by a click.
    for question in questions:
        choices[question] = "0"
    # Remove the onboarding screen and add a questions frame with the next and previous buttons
    onboard.grid_forget()
    q_frame = Frame(root)
    q_frame.grid(row=1, column=1)
    q_label = Label(q_frame, text=str(index + 1) + ". " + questions[index])
    q_label.grid(row=0, column=1)
    display_options(questions[index])
    prv_btn = Button(q_frame, text="Previous",
                     state=DISABLED, bg="#2e6930", fg="#fff")
    prv_btn.grid(row=3, column=0, padx=10, pady=10)
    nxt_btn = Button(q_frame, text="Next", command=nxt_que,
                     bg="#2e6930", fg="#fff")
    nxt_btn.grid(row=3, column=2, padx=10, pady=10)
    run_timer()

# --------------------------------Function to display the questions


def display_questions(index):
    global q_frame
    global q_label
    global prv_btn
    global nxt_btn

    q_frame.grid_forget()
    q_label.grid_forget()

    q_frame = Frame(root)
    q_frame.grid(row=1, column=1)
    q_label = Label(q_frame, text=str(index + 1) + ". " + questions[index])
    q_label.grid(row=0, column=1)
# -------------------------------Function to display the options to the questions


def display_options(question):
    global test
    global q_frame
    global var
    global choices

    option_frame = Frame(q_frame)
    option_frame.grid(row=1, column=1)

    options = test[question]
    var = StringVar()
    var.set(choices[question])
    random.shuffle(options[0:len(options)-1])
    for option in options[0:len(options)-1]:
        Radiobutton(option_frame, text=option,
                    variable=var, value=option).pack()


# Widgets
# -----------------------------------------------------------------------TITLE OF TEST
title_label = Label(
    root, text=active_sheet["B1"].value, font=("Helvetica", 10))
title_label.grid(row=0, column=1, padx=50, pady=(20, 30))

# ------------------------------------------------------------------------STUDENT DETAILS
onboard = LabelFrame(root)
onboard.grid(row=1, column=1, padx=50, pady=(20, 50), sticky=W+E)

name_label = Label(onboard, text="Name: ")
name_label.grid(row=0, column=0, padx=5, pady=10)
name_field = Entry(onboard, bd=3, width=20)
name_field.grid(row=0, column=1, padx=5, pady=10)

matric_label = Label(onboard, text="Reg No.: ")
matric_label.grid(row=1, column=0, padx=5, pady=10)
matric_field = Entry(onboard, bd=3, width=20)
matric_field.grid(row=1, column=1, padx=5, pady=10)

level_label = Label(onboard, text="Level: ")
level_label.grid(row=2, column=0, padx=5, pady=10)
level_field = Entry(onboard, bd=3, width=20)
level_field.grid(row=2, column=1, padx=5, pady=10)
# -------------------------------------------------------------------TAKE TEST BUTTON
take_button = Button(onboard, text="Take Test",
                     command=take_test, bg="#2e6930", fg="#fff")
take_button.grid(row=3, column=0, padx=10, pady=10,
                 columnspan=2, ipadx=5, ipady=5, sticky=W)

submit_btn = None

root.mainloop()
